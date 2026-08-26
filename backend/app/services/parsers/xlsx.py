# Excel(.xlsx) 解析：每个工作表转成"每行单元格竖线分隔"的文本。
# 这样切片检索命中时，能大致保留表格的行结构。
from pathlib import Path

from openpyxl import load_workbook

from app.services.parsers.base import ParseError, Segment

# 单个工作表最多读取的行数，防止超大表格拖垮索引
MAX_ROWS_PER_SHEET = 3000


def parse(path: str | Path) -> list[Segment]:
    try:
        # read_only=True 流式读取省内存；data_only=True 取公式的计算结果而非公式文本
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as e:
        raise ParseError(f"Excel 文件无法解析: {e}")

    segments: list[Segment] = []
    for sheet in wb.worksheets:
        lines: list[str] = []
        truncated = False
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= MAX_ROWS_PER_SHEET:
                truncated = True
                break
            # 单元格转字符串；换行替换为空格，避免破坏行结构
            cells = ["" if v is None else str(v).replace("\n", " ").strip() for v in row]
            line = " | ".join(cells).strip(" |")
            if line:
                lines.append(line)
        if not lines:
            continue
        text = "\n".join(lines)
        if truncated:
            text += f"\n（表格过长，仅保留前 {MAX_ROWS_PER_SHEET} 行）"
        # 工作表名作为出处元信息，引用时展示"工作表「Sheet1」"
        segments.append(Segment(text=text, meta={"sheet": sheet.title}))
    wb.close()

    if not segments:
        raise ParseError("Excel 中没有可提取的内容")
    return segments
